from openpyxl import load_workbook
from pprint import pprint
import re

ROUTE_COL = 2
MATRIX_COL = 3

BOUND_ROW = 1
DAYS_ROW = 2
MATRIX_ROW = 4

# Index into a sheet using row and col
def idx(row, col):
  return chr(ord('A')+col-1)+str(row)

# Test string for a value indicating end of processing
# Should eventually just be the test route (9900)
def is_end(s):
  return s != None and s.startswith("9900")

# Load the workbook as readonly to save memory space 
def load_schedule():
  wb = load_workbook(filename='tcat-schedule.xlsx')
  ws = wb["schedule"]
  return ws

def next_matrix(ws, row):
  # Route number and name
  route = ws[idx(row, ROUTE_COL)].value.strip()
  # Route bound: Loop, Outbound, Inbound
  bound = ws[idx(row+BOUND_ROW, ROUTE_COL)].value
  # Days that this route runs
  days = ws[idx(row+DAYS_ROW, ROUTE_COL)].value

  row += MATRIX_ROW
  num_cols = 0

  # List of stops for
  stops = []
  while ws[idx(row, MATRIX_COL+num_cols)].value != None:
    stops.append(ws[idx(row, MATRIX_COL+num_cols)].value)
    num_cols += 1

  # The schedule for this route-bound-days tuple
  matrix = []
  row += 1
  while ws[idx(row, MATRIX_COL)].value != None:
    col = 0
    data = []
    while col < num_cols:
      data.append(ws[idx(row, MATRIX_COL+col)].value)
      col += 1
    matrix.append(data)
    row += 1

  row_value = ws[idx(row, ROUTE_COL)].value
  while True:
    if row_value != None:
      row_value = row_value.strip()
      if re.match('(\\d+)\\s+.*', row_value) != None:
        break;
    row += 1
    row_value = ws[idx(row, ROUTE_COL)].value

  return (route, bound, days, stops, matrix, row)

# Parse the next route from the file
# Requires: value at (row, ROUTE_COL) be the the start of a new schedule
def next_route(ws, row):
  current_route = []
  next_row = 0
  (route, bound, days, stops, matrix, next_row) = next_matrix(ws, row)
  current_route.append((route, bound, days, stops, matrix))
  next_route = ws[idx(next_row, ROUTE_COL)].value.strip()
  while route == next_route:
    row = next_row
    (route, bound, days, stops, matrix, next_row) = next_matrix(ws, row)
    current_route.append((route, bound, days, stops, matrix))
    next_route = ws[idx(next_row, ROUTE_COL)].value.strip()
  return (current_route, next_row)


def merge_matrices(route_data):
  n = len(route_data)
  if n == 1:
    (route, bound, days, stops, matrix) = route_data[0]
    return (route, [(days, bound, stops, matrix)])
  outbound = route_data[:n//2]
  inbound =route_data[n//2:]
  n = n//2

  route = ""
  updated_route_data = []
  for i in range(n):
    (route1, bound1, days1, stops1, matrix1) = outbound[i]
    (route2, bound2, days2, stops2, matrix2) = inbound[i]
    route = route1
    days = days1
    bounds = [
      (bound1, (0, len(stops1))),
      (bound2, (len(stops1), len(stops1) + len(stops2)))
    ]
    stops = stops1 + stops2
    matrix = []
    
    rows = min(len(matrix1), len(matrix2))
    for j in range(rows):
      #print(j)
      matrix.append(matrix1[j] + matrix2[j])
    if len(matrix1) > len(matrix2):
      matrix.append(matrix1[-1])
    else:
      matrix.append(matrix2[-1])

    updated_route_data.append((days, bounds, stops, matrix))
  return (route, updated_route_data)

def load_routes(ws):
  row = 1
  tcat_data = []
  while not is_end(ws[idx(row, ROUTE_COL)].value):
    (route_data, next_row) = next_route(ws, row)
    merged_data = merge_matrices(route_data)
    tcat_data.append(merged_data)
    row = next_row
  return tcat_data

